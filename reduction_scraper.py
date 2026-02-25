#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股减持公告自动化抓取与解析系统
Stock Reduction Announcement Daily Reporter

功能概述:
  1. 每日 21:00 自动从巨潮资讯网(cninfo.com.cn) 获取全 A 股减持类公告
  2. 正则表达式从公告标题 + 摘要中提取关键业务字段
  3. 融合 AKShare 实时行情数据（总市值、最新收盘价）
  4. 输出格式化 Excel 报表，含汇总统计 Sheet

安装依赖 (install dependencies):
  pip install akshare pandas requests openpyxl schedule pdfplumber beautifulsoup4 lxml

可选增强 (optional):
  - pdfplumber: 解析公告 PDF 正文，提取更完整的字段
  - beautifulsoup4 + lxml: 解析 HTML 摘要

字段提取逻辑说明 (见底部文档块):
  详见模块 __doc__ 末尾的 "字段提取策略" 说明。

运行方式:
  python reduction_scraper.py --mode run          # 立即执行今日
  python reduction_scraper.py --mode run --date 2025-01-15  # 指定日期
  python reduction_scraper.py --mode daemon       # 守护进程（每日21:00）
  python reduction_scraper.py --mode test         # 测试模式

作者: AI Financial Data Engineer
版本: 2.0.0
"""

# ============================================================
# § 1. 模块导入与全局配置
# ============================================================

import sys
import os
import re
import json
import time
import logging
import warnings
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple, Any

import requests
import pandas as pd
import schedule

warnings.filterwarnings("ignore")

# ─── 全局配置字典 ─────────────────────────────────────────────────────────────
CONFIG: Dict[str, Any] = {
    # 输入/输出
    "output_dir": "./reports",                  # Excel 报表输出目录
    "log_dir": "./logs",                        # 日志目录

    # 定时任务
    "schedule_time": "21:00",                   # 每日定时执行时间

    # 网络请求
    "request_delay": 1.5,                       # 请求间隔秒数（防反爬）
    "request_timeout": 30,                      # 单次请求超时秒数
    "max_pages": 15,                            # 最大翻页数（每页30条）
    "max_retries": 3,                           # 请求失败最大重试次数

    # 行情数据
    "market_data_delay": 3,                     # 获取行情前的等待秒数

    # ── 减持公告过滤关键词 ──────────────────────────────────────────────────
    # 标题必须包含下列至少一个词才纳入分析
    "reduction_keywords": [
        "拟减持", "减持股份计划", "减持进展", "减持计划",
        "减持方案", "减持公告", "集中竞价减持", "大宗交易减持",
        "减持股份", "减少持股比例", "预减持", "减持预披露",
        "减持数量", "股份减持", "股东减持",
    ],

    # 若标题以下列词为"主语"则排除（例如纯增持公告误判）
    "title_exclude_patterns": [
        r"^关于.*增持.*的公告$",        # 纯增持公告
        r"^.*股票回购.*$",              # 回购公告
    ],
}


# ============================================================
# § 2. 日志系统初始化
# ============================================================

def _setup_logging() -> logging.Logger:
    """初始化双通道日志（控制台 + 月度滚动日志文件）"""
    log_dir = Path(CONFIG["log_dir"])
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"reduction_{datetime.now():%Y%m}.log"

    fmt = "%(asctime)s [%(levelname)-5s] %(name)s:%(lineno)d — %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"

    root = logging.getLogger()
    root.setLevel(logging.INFO)
    if not root.handlers:
        root.addHandler(
            logging.FileHandler(log_file, encoding="utf-8")
        )
        root.addHandler(logging.StreamHandler(sys.stdout))

    for h in root.handlers:
        h.setFormatter(logging.Formatter(fmt, datefmt))

    return logging.getLogger("reduction_scraper")


logger = _setup_logging()


# ============================================================
# § 3. 工具函数
# ============================================================

def build_cninfo_url(adjunct_url: str) -> str:
    """
    根据巨潮资讯 API 返回的 adjunctUrl 字段构建 PDF 直达链接。

    CNINFO adjunctUrl 格式样例:
        "2024-01-15/1234567890abcdef.PDF"
    最终 URL:
        "http://static.cninfo.com.cn/finalpage/2024-01-15/1234567890abcdef.PDF"
    """
    if adjunct_url and adjunct_url.strip():
        return f"http://static.cninfo.com.cn/finalpage/{adjunct_url.strip()}"
    return "N/A"


def ms_timestamp_to_date(ts_ms: Any) -> str:
    """毫秒时间戳 → YYYY-MM-DD 字符串，转换失败返回 'N/A'"""
    try:
        return datetime.fromtimestamp(int(ts_ms) / 1000).strftime("%Y-%m-%d")
    except Exception:
        return "N/A"


def retry_request(
    session: requests.Session,
    method: str,
    url: str,
    max_retries: int = 3,
    **kwargs,
) -> Optional[requests.Response]:
    """带指数退避重试的 HTTP 请求包装器"""
    for attempt in range(1, max_retries + 1):
        try:
            resp = getattr(session, method)(url, timeout=CONFIG["request_timeout"], **kwargs)
            resp.raise_for_status()
            return resp
        except requests.exceptions.RequestException as exc:
            wait = 2 ** attempt
            logger.warning(f"请求失败 (第{attempt}次/{max_retries}): {exc}，{wait}s 后重试")
            if attempt < max_retries:
                time.sleep(wait)
    return None


# ============================================================
# § 4. 数据获取模块 —— 巨潮资讯 CNINFO Fetcher
# ============================================================

class CNInfoFetcher:
    """
    巨潮资讯网公告获取器。

    核心 API 端点:
        POST http://www.cninfo.com.cn/new/hisAnnouncement/query

    该接口为巨潮资讯网前端使用的公开接口，无需登录/鉴权。
    每页最多返回 30 条，通过 pageNum 翻页。
    """

    QUERY_URL = "http://www.cninfo.com.cn/new/hisAnnouncement/query"

    _HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Referer": "http://www.cninfo.com.cn/new/commonUrl/pageOfSearch",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Origin": "http://www.cninfo.com.cn",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(self._HEADERS)

    # ── 公开方法 ───────────────────────────────────────────────────────────

    def fetch_reduction_announcements(self, query_date: str) -> List[Dict]:
        """
        获取指定日期全市场减持相关公告（沪深两市）。

        实现策略:
          - 使用两个搜索词 ["减持股份", "减持计划"] 分别查询，最大化召回率
          - 按 announcementId 去重
          - 交给 filter_by_title() 做精确过滤

        Args:
            query_date: 'YYYY-MM-DD'

        Returns:
            原始公告字典列表（未过滤）
        """
        date_range = f"{query_date}~{query_date}"
        seen_ids: set = set()
        all_ann: List[Dict] = []

        # 多关键词搜索提升召回
        search_terms = ["减持股份", "减持计划", "拟减持"]
        for term in search_terms:
            logger.info(f"  ↳ 搜索词: 「{term}」 | 日期范围: {date_range}")
            batch = self._paginated_query(term, date_range)
            for ann in batch:
                ann_id = ann.get("announcementId", "")
                if ann_id and ann_id not in seen_ids:
                    seen_ids.add(ann_id)
                    all_ann.append(ann)
            time.sleep(CONFIG["request_delay"])

        logger.info(f"原始公告（去重后）共 {len(all_ann)} 条")
        return all_ann

    def filter_by_title(self, announcements: List[Dict]) -> List[Dict]:
        """
        按标题关键词精确过滤减持公告。

        保留规则:
          - 标题含 CONFIG['reduction_keywords'] 中至少一个词
          - 标题不匹配 CONFIG['title_exclude_patterns'] 中任一正则

        Returns:
            过滤后的减持公告列表
        """
        result = []
        for ann in announcements:
            title = ann.get("announcementTitle", "")
            if self._is_reduction_title(title):
                result.append(ann)
        logger.info(f"标题过滤后，减持公告 {len(result)} 条")
        return result

    # ── 私有辅助方法 ──────────────────────────────────────────────────────

    def _paginated_query(self, keyword: str, date_range: str) -> List[Dict]:
        """分页拉取指定关键词的公告列表"""
        items: List[Dict] = []
        for page in range(1, CONFIG["max_pages"] + 1):
            payload = {
                "pageNum": page,
                "pageSize": 30,
                "tabName": "fulltext",      # 全文检索模式
                "column": "",               # 空 = 全市场（沪 + 深）
                "category": "",
                "platId": "",
                "seDate": date_range,
                "searchkey": keyword,
                "secid": "",
                "stock": "",
                "isHLtitle": "true",        # 高亮匹配标题词
            }
            resp = retry_request(
                self.session, "post", self.QUERY_URL,
                data=payload,
                max_retries=CONFIG["max_retries"],
            )
            if resp is None:
                logger.warning(f"  第{page}页请求失败，终止翻页")
                break

            try:
                data = resp.json()
            except json.JSONDecodeError:
                logger.warning(f"  第{page}页 JSON 解析失败")
                break

            batch = data.get("announcements") or []
            if not batch:
                break
            items.extend(batch)

            total = int(data.get("totalRecordNum", 0))
            logger.debug(f"  第{page}页 {len(batch)} 条 | 总计 {total} 条")
            if page * 30 >= total:
                break
            time.sleep(CONFIG["request_delay"])

        return items

    @staticmethod
    def _is_reduction_title(title: str) -> bool:
        """判断标题是否属于减持公告"""
        has_keyword = any(kw in title for kw in CONFIG["reduction_keywords"])
        if not has_keyword:
            return False
        for pat in CONFIG["title_exclude_patterns"]:
            if re.match(pat, title):
                return False
        return True


# ============================================================
# § 5. 文本解析模块 —— 正则字段提取器
# ============================================================

class ReductionFieldExtractor:
    """
    从公告标题（及可选摘要/正文）中用正则提取减持关键字段。

    ┌─────────────────────────────────────────────────────────────────┐
    │                    字段提取策略说明                              │
    ├──────────────┬──────────────────────────────────────────────────┤
    │ 字段         │ 提取策略                                         │
    ├──────────────┼──────────────────────────────────────────────────┤
    │ 股东主体     │ 正则捕获"关于X拟减持"/"X持股5%以上股东"等前置主语│
    │ 减持比例     │ 匹配"不超过总股本X%"/"占总股本X%"等数字+%模式    │
    │ 减持数量     │ 匹配"不超过X万股"/"X股"等数量模式，统一为万股    │
    │ 减持方式     │ 关键词字典匹配（大宗/竞价/协议/定向等）          │
    │ 股份来源     │ 关键词字典匹配（IPO前/定增/激励/员工持股等）     │
    └──────────────┴──────────────────────────────────────────────────┘

    无法提取 → 填充 "需人工核实"，绝不抛异常中断流程。
    """

    # ── 减持比例：捕获"总股本的 X%"中的 X ─────────────────────────────
    _RATIO_PATTERNS: List[str] = [
        r"不超过.*?总股本(?:的|比例约?|比)?[\s：:]*(\d+\.?\d*)\s*%",
        r"占.*?总股本.*?(\d+\.?\d*)\s*%",
        r"总股本比例.*?(\d+\.?\d*)\s*%",
        r"减持比例.*?(\d+\.?\d*)\s*%",
        r"(\d+\.?\d*)\s*%.*?总股本",
        r"合计不超过.*?(\d+\.?\d*)\s*%",
    ]

    # ── 减持数量：捕获万股数量 ─────────────────────────────────────────
    _QTY_PATTERNS: List[Tuple[str, str]] = [
        # (pattern, unit)  unit='wan' | 'share'
        (r"不超过\s*(\d[\d,，]*(?:\.\d+)?)\s*万\s*股",   "wan"),
        (r"合计不超过\s*(\d[\d,，]*(?:\.\d+)?)\s*万\s*股", "wan"),
        (r"减持\s*(\d[\d,，]*(?:\.\d+)?)\s*万\s*股",      "wan"),
        (r"不超过\s*(\d[\d,，]*)\s*股",                   "share"),
        (r"(\d[\d,，]*(?:\.\d+)?)\s*万股",                "wan"),
    ]

    # ── 减持方式：关键词到规范化名称 ──────────────────────────────────
    _METHOD_MAP: Dict[str, List[str]] = {
        "大宗交易":   ["大宗交易"],
        "集中竞价":   ["集中竞价", "二级市场竞价"],
        "协议转让":   ["协议转让", "协议方式"],
        "询价转让":   ["询价转让"],
        "定向转让":   ["定向转让", "非交易过户"],
        "二级市场":   ["二级市场"],
    }

    # ── 股份来源：关键词到规范化名称 ──────────────────────────────────
    _SOURCE_MAP: Dict[str, List[str]] = {
        "IPO 前取得":    ["首次公开发行前", "上市前", "IPO前", "发行前持有", "原始股"],
        "定向增发取得":  ["定向增发", "非公开发行", "定增取得"],
        "股权激励取得":  ["股权激励", "限制性股票", "股票期权", "激励计划授予"],
        "员工持股计划":  ["员工持股计划", "ESOP"],
        "大宗交易受让":  ["大宗交易受让", "受让取得", "受让股份"],
        "其他":          ["继承", "赠与", "司法拍卖", "竞拍"],
    }

    # ── 股东主体提取模式（按优先级排序）────────────────────────────────
    _HOLDER_PATTERNS: List[str] = [
        # 格式1: 关于【股东名称】拟减持…
        r"关于([\u4e00-\u9fa5A-Za-z0-9（()）\-·%\.]+?)(?:拟|计划|将)(?:通过)?(?:大宗|集中|协议|询价|二级|减持|出售)",
        # 格式2: 关于【股东名称】持股5%以上股东…
        r"关于([\u4e00-\u9fa5A-Za-z0-9（()）\-·%\.]+?)(?:持股\d+%以上|控股|实控人|董事|监事|高管)",
        # 格式3: 【公司类主体】拟减持（含有限公司/合伙企业/基金）
        r"([\u4e00-\u9fa5A-Za-z0-9（()）\-·]+?(?:有限公司|股份有限公司|合伙企业|基金|中心))(?:.*?)(?:拟|计划)减持",
        # 格式4: 自然人姓名（2-4 个汉字 + 先生/女士）
        r"([\u4e00-\u9fa5]{2,4}(?:先生|女士)).*?(?:拟|计划)减持",
        # 格式5: 兜底——标题前半段（"关于"后至"拟"前）
        r"关于([\u4e00-\u9fa5A-Za-z0-9（()）\-·%]{2,20}?)拟",
    ]

    # ── 公开方法 ───────────────────────────────────────────────────────

    def extract(self, announcement: Dict) -> Dict:
        """
        从一条公告 dict 中提取所有业务字段，不抛异常。

        Args:
            announcement: CNINFO API 返回的原始公告字典

        Returns:
            结构化字段字典
        """
        title = announcement.get("announcementTitle", "")
        sec_code = str(announcement.get("secCode", "")).zfill(6)
        sec_name = announcement.get("secName", "")
        adjunct_url = announcement.get("adjunctUrl", "")
        ann_time = announcement.get("announcementTime", 0)

        record = {
            "公告日期":          ms_timestamp_to_date(ann_time),
            "股票代码":          sec_code,
            "股票简称":          sec_name,
            "减持股东":          self._safe(self._extract_holder, title),
            "拟减持比例(%)":     self._safe(self._extract_ratio, title),
            "拟减持数量(万股)":  self._safe(self._extract_quantity, title),
            "减持方式":          self._safe(self._extract_methods, title),
            "股份来源":          self._safe(self._extract_source, title),
            "公告标题":          title,
            "公告原文链接":      build_cninfo_url(adjunct_url),
        }
        return record

    # ── 私有提取方法 ──────────────────────────────────────────────────

    @staticmethod
    def _safe(func, text: str) -> str:
        """包装提取函数，任何异常均返回 '需人工核实'"""
        try:
            result = func(text)
            return result if result else "需人工核实"
        except Exception:
            return "需人工核实"

    def _extract_ratio(self, text: str) -> str:
        for pat in self._RATIO_PATTERNS:
            m = re.search(pat, text)
            if m:
                val = m.group(1).strip()
                try:
                    float(val)      # 校验是合法数字
                    return val
                except ValueError:
                    continue
        return ""

    def _extract_quantity(self, text: str) -> str:
        for pat, unit in self._QTY_PATTERNS:
            m = re.search(pat, text)
            if m:
                raw = m.group(1).replace(",", "").replace("，", "")
                try:
                    qty = float(raw)
                    if unit == "share":
                        qty = qty / 10000   # 转为万股
                    return f"{qty:.4f}".rstrip("0").rstrip(".")
                except ValueError:
                    continue
        return ""

    def _extract_methods(self, text: str) -> str:
        found = [
            name for name, kws in self._METHOD_MAP.items()
            if any(kw in text for kw in kws)
        ]
        return "、".join(found) if found else ""

    def _extract_source(self, text: str) -> str:
        found = [
            name for name, kws in self._SOURCE_MAP.items()
            if any(kw in text for kw in kws)
        ]
        return "、".join(found) if found else ""

    def _extract_holder(self, text: str) -> str:
        for pat in self._HOLDER_PATTERNS:
            m = re.search(pat, text)
            if m:
                candidate = m.group(1).strip()
                # 过滤掉太短（≤1字）或太长（>20字）的噪音匹配
                if 2 <= len(candidate) <= 25:
                    return candidate
        return ""


# ============================================================
# § 6. 可选增强：PDF 正文解析（需安装 pdfplumber）
# ============================================================

class PDFEnhancer:
    """
    可选模块：下载公告 PDF 并用 pdfplumber 解析正文，
    补充从标题无法提取的字段（股份来源、精确减持比例等）。

    使用前需安装:
        pip install pdfplumber

    调用示例:
        enhancer = PDFEnhancer()
        extra = enhancer.enhance_record(record)
        record.update(extra)
    """

    _available: bool = False

    def __init__(self):
        try:
            import pdfplumber          # noqa: F401
            self._available = True
            logger.info("pdfplumber 可用，将启用 PDF 正文增强解析")
        except ImportError:
            logger.info("pdfplumber 未安装，跳过 PDF 增强解析（pip install pdfplumber 可启用）")

    def enhance_record(self, record: Dict) -> Dict:
        """
        尝试从 PDF 中补充字段，返回增量字段 dict。
        若 pdfplumber 不可用或下载失败，返回空 dict 不影响主流程。
        """
        if not self._available:
            return {}

        url = record.get("公告原文链接", "")
        if not url or url == "N/A" or not url.lower().endswith(".pdf"):
            return {}

        try:
            import pdfplumber
            import io

            resp = requests.get(url, timeout=20, headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": "http://www.cninfo.com.cn/",
            })
            resp.raise_for_status()

            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                # 只取前 3 页，减少解析开销
                text = "\n".join(
                    (page.extract_text() or "") for page in pdf.pages[:3]
                )

            return self._parse_pdf_text(text, record)

        except Exception as e:
            logger.debug(f"PDF 增强解析失败 [{url}]: {e}")
            return {}

    @staticmethod
    def _parse_pdf_text(text: str, existing: Dict) -> Dict:
        """从 PDF 文本中补充缺失字段"""
        updates: Dict[str, str] = {}
        extractor = ReductionFieldExtractor()

        # 仅在标题层未能提取时，用 PDF 文本补充
        if existing.get("拟减持比例(%)") == "需人工核实":
            ratio = extractor._safe(extractor._extract_ratio, text)
            if ratio != "需人工核实":
                updates["拟减持比例(%)"] = ratio

        if existing.get("拟减持数量(万股)") == "需人工核实":
            qty = extractor._safe(extractor._extract_quantity, text)
            if qty != "需人工核实":
                updates["拟减持数量(万股)"] = qty

        if existing.get("减持方式") == "需人工核实":
            method = extractor._safe(extractor._extract_methods, text)
            if method != "需人工核实":
                updates["减持方式"] = method

        if existing.get("股份来源") == "需人工核实":
            source = extractor._safe(extractor._extract_source, text)
            if source != "需人工核实":
                updates["股份来源"] = source

        if existing.get("减持股东") == "需人工核实":
            holder = extractor._safe(extractor._extract_holder, text)
            if holder != "需人工核实":
                updates["减持股东"] = holder

        if updates:
            logger.debug(f"PDF 增强补充字段: {list(updates.keys())}")
        return updates


# ============================================================
# § 7. 行情数据模块 —— AKShare Market Data
# ============================================================

class MarketDataFetcher:
    """
    从 AKShare 获取 A 股实时行情，重点提取总市值与最新收盘价。

    核心接口: ak.stock_zh_a_spot_em()
    返回字段（部分）: 代码, 名称, 最新价, 总市值, 流通市值, ...
    """

    _CACHE_TTL_SECONDS = 1800           # 缓存有效期 30 分钟

    def __init__(self):
        self._cache_df: Optional[pd.DataFrame] = None
        self._cache_ts: Optional[float] = None

    def get_market_cap(self, stock_codes: List[str]) -> pd.DataFrame:
        """
        批量查询股票总市值与最新价。

        Args:
            stock_codes: 股票代码列表（6位字符串）

        Returns:
            DataFrame，列: 股票代码, 最新收盘价, 总市值(亿元), 流通市值(亿元)
        """
        all_df = self._load_all_market_data()
        if all_df.empty:
            return pd.DataFrame()

        normalized = [str(c).zfill(6) for c in stock_codes]
        mask = all_df["股票代码"].isin(normalized)
        return all_df.loc[mask, ["股票代码", "最新收盘价", "总市值(亿元)", "流通市值(亿元)"]].copy()

    def _load_all_market_data(self) -> pd.DataFrame:
        """带缓存的全市场行情加载"""
        now = time.time()
        if (
            self._cache_df is not None
            and self._cache_ts is not None
            and (now - self._cache_ts) < self._CACHE_TTL_SECONDS
        ):
            logger.info("  ↳ 使用行情缓存数据（缓存未过期）")
            return self._cache_df

        try:
            import akshare as ak
            logger.info("  ↳ 调用 AKShare stock_zh_a_spot_em() 获取全市场行情...")
            raw = ak.stock_zh_a_spot_em()

            # 统一列名
            col_map = {"代码": "股票代码", "名称": "股票简称_AK", "最新价": "最新收盘价",
                       "总市值": "总市值(元)", "流通市值": "流通市值(元)"}
            df = raw.rename(columns=col_map)
            df = df[list(col_map.values())].copy()

            df["股票代码"] = df["股票代码"].astype(str).str.zfill(6)
            df["总市值(亿元)"] = (pd.to_numeric(df["总市值(元)"], errors="coerce") / 1e8).round(2)
            df["流通市值(亿元)"] = (pd.to_numeric(df["流通市值(元)"], errors="coerce") / 1e8).round(2)
            df["最新收盘价"] = pd.to_numeric(df["最新收盘价"], errors="coerce")

            # 仅保留后续需要的列
            df = df[["股票代码", "最新收盘价", "总市值(亿元)", "流通市值(亿元)"]]

            self._cache_df = df
            self._cache_ts = now
            logger.info(f"  ↳ 行情数据加载完成，共 {len(df)} 只股票")
            return df

        except Exception as exc:
            logger.error(f"AKShare 行情数据获取失败: {exc}", exc_info=True)
            return pd.DataFrame()


# ============================================================
# § 8. 报表生成模块 —— Excel 输出
# ============================================================

class ExcelReporter:
    """
    将结构化减持数据与行情数据合并，导出为带格式的 Excel 报表。

    Sheet 结构:
        - 减持公告明细: 按总市值降序排列的主数据表
        - 每日统计摘要: 公告数量/类型分布等汇总指标
    """

    # 最终输出列顺序
    _COLUMN_ORDER = [
        "公告日期", "股票代码", "股票简称",
        "减持股东", "拟减持比例(%)", "拟减持数量(万股)",
        "减持方式", "股份来源",
        "最新收盘价", "总市值(亿元)", "流通市值(亿元)",
        "公告标题", "公告原文链接",
    ]

    def __init__(self, output_dir: str = CONFIG["output_dir"]):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, records: List[Dict], query_date: str) -> str:
        """
        生成 Excel 报表。

        Args:
            records:    已提取字段的公告记录列表
            query_date: 查询日期 'YYYY-MM-DD'

        Returns:
            输出文件的绝对路径字符串；无数据时返回空字符串
        """
        if not records:
            logger.warning("records 为空，跳过报表生成")
            return ""

        df = pd.DataFrame(records)
        df = self._merge_market_data(df)
        df = self._clean_and_sort(df)
        df = self._reorder_columns(df)

        date_str = query_date.replace("-", "")
        out_path = self.output_dir / f"Reduction_Daily_Report_{date_str}.xlsx"
        self._write_excel(df, str(out_path), query_date)

        logger.info(f"报表已生成 → {out_path}")
        return str(out_path)

    # ── 私有方法 ──────────────────────────────────────────────────────────

    def _merge_market_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """合并 AKShare 行情数据"""
        codes = df["股票代码"].unique().tolist()
        logger.info(f"  ↳ 获取 {len(codes)} 只股票的行情数据...")
        time.sleep(CONFIG["market_data_delay"])

        fetcher = MarketDataFetcher()
        mkt_df = fetcher.get_market_cap(codes)

        if not mkt_df.empty:
            df = df.merge(mkt_df, on="股票代码", how="left")
            logger.info(f"  ↳ 行情合并完成，匹配率 {mkt_df['股票代码'].isin(codes).sum()}/{len(codes)}")
        else:
            for col in ["最新收盘价", "总市值(亿元)", "流通市值(亿元)"]:
                df[col] = None
            logger.warning("  ↳ 行情数据为空，市值列填 None")
        return df

    def _clean_and_sort(self, df: pd.DataFrame) -> pd.DataFrame:
        """数值化并排序（总市值降序 → 减持比例降序）"""
        df["_mkt_cap"] = pd.to_numeric(df.get("总市值(亿元)"), errors="coerce")
        df["_ratio"]   = pd.to_numeric(df.get("拟减持比例(%)"), errors="coerce")

        sort_keys = [k for k in ["_mkt_cap", "_ratio"] if k in df.columns]
        if sort_keys:
            df = df.sort_values(sort_keys, ascending=False, na_position="last")

        df = df.drop(columns=["_mkt_cap", "_ratio"], errors="ignore")
        return df.reset_index(drop=True)

    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        ordered = [c for c in self._COLUMN_ORDER if c in df.columns]
        extra = [c for c in df.columns if c not in ordered]
        return df[ordered + extra]

    def _write_excel(self, df: pd.DataFrame, path: str, query_date: str):
        """写入 Excel 并应用格式样式"""
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="减持公告明细", index=False)
            self._add_summary_sheet(writer, df, query_date)
            self._apply_styles(writer, df)

    def _apply_styles(self, writer, df: pd.DataFrame):
        """应用表格格式: 冻结行、标题色、自动列宽、超链接、间行底色"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl 样式模块不可用，跳过格式化")
            return

        ws = writer.sheets["减持公告明细"]

        # 冻结首行
        ws.freeze_panes = "A2"

        # 标题行：深蓝底白字
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 找出"公告原文链接"所在列索引
        link_col_idx: Optional[int] = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == "公告原文链接":
                link_col_idx = idx
                break

        # 超链接 + 需人工核实高亮
        manual_fill  = PatternFill("solid", fgColor="FFEB9C")
        link_font    = Font(color="0563C1", underline="single", size=9)
        alt_row_fill = PatternFill("solid", fgColor="EBF3FB")
        normal_font  = Font(size=9)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = normal_font
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if str(cell.value) == "需人工核实":
                    cell.fill = manual_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_row_fill

            # 超链接处理
            if link_col_idx:
                link_cell = ws.cell(row=row_idx, column=link_col_idx)
                url_val = str(link_cell.value or "")
                if url_val.startswith("http"):
                    link_cell.hyperlink = url_val
                    link_cell.value = "查看原文↗"
                    link_cell.font = link_font

        # 自动列宽（中文字符按1.8宽度估算）
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            try:
                max_len = max(
                    len(str(col_name)),
                    int(df.iloc[:, col_idx - 1].astype(str).str.len().quantile(0.9)),
                )
            except Exception:
                max_len = len(str(col_name))
            ws.column_dimensions[col_letter].width = min(max_len * 1.8 + 2, 55)

        # 行高
        for row_idx in range(1, len(df) + 2):
            ws.row_dimensions[row_idx].height = 16

    def _add_summary_sheet(self, writer, df: pd.DataFrame, query_date: str):
        """添加每日统计摘要 Sheet"""

        def _count_contains(col: str, keyword: str) -> int:
            if col in df.columns:
                return int(df[col].astype(str).str.contains(keyword, na=False).sum())
            return 0

        def _to_num(series) -> pd.Series:
            return pd.to_numeric(series, errors="coerce")

        ratio_series = _to_num(
            df["拟减持比例(%)"].replace("需人工核实", None)
        ) if "拟减持比例(%)" in df.columns else pd.Series(dtype=float)

        stats = {
            "统计项目": [
                "报告日期", "减持公告总数", "涉及上市公司数",
                "已明确减持比例条数", "含大宗交易条数", "含集中竞价条数",
                "含协议转让条数", "最大减持比例(%)", "平均减持比例(%)",
                "最大总市值(亿元)", "报告生成时间",
            ],
            "数值": [
                query_date,
                len(df),
                df["股票代码"].nunique() if "股票代码" in df.columns else "N/A",
                int((df["拟减持比例(%)"] != "需人工核实").sum()) if "拟减持比例(%)" in df.columns else "N/A",
                _count_contains("减持方式", "大宗交易"),
                _count_contains("减持方式", "集中竞价"),
                _count_contains("减持方式", "协议转让"),
                round(ratio_series.max(), 2) if not ratio_series.empty else "N/A",
                round(ratio_series.mean(), 2) if not ratio_series.empty else "N/A",
                df["总市值(亿元)"].max() if "总市值(亿元)" in df.columns else "N/A",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        }
        summary_df = pd.DataFrame(stats)
        summary_df.to_excel(writer, sheet_name="每日统计摘要", index=False)


# ============================================================
# § 9. 主流程编排 —— Pipeline Orchestrator
# ============================================================

class ReductionPipeline:
    """
    减持公告数据管道主控制器。

    执行顺序:
        Step 1 → CNInfoFetcher.fetch_reduction_announcements()
        Step 2 → CNInfoFetcher.filter_by_title()
        Step 3 → ReductionFieldExtractor.extract()  [+ PDFEnhancer 可选]
        Step 4 → ExcelReporter.generate()
    """

    def __init__(self, enable_pdf_enhance: bool = False):
        self.fetcher   = CNInfoFetcher()
        self.extractor = ReductionFieldExtractor()
        self.enhancer  = PDFEnhancer() if enable_pdf_enhance else None
        self.reporter  = ExcelReporter()

    def run(self, query_date: Optional[str] = None) -> Optional[str]:
        """
        执行完整数据管道。

        Args:
            query_date: 'YYYY-MM-DD' 格式日期，默认为当天

        Returns:
            Excel 文件路径，失败或无数据时返回 None
        """
        if query_date is None:
            query_date = datetime.now().strftime("%Y-%m-%d")

        _banner = "=" * 65
        logger.info(_banner)
        logger.info(f"  A股减持公告自动化系统 启动 | 查询日期: {query_date}")
        logger.info(_banner)

        try:
            # ── Step 1: 获取公告 ────────────────────────────────────────
            logger.info("[1/4] 从巨潮资讯网拉取公告数据...")
            raw = self.fetcher.fetch_reduction_announcements(query_date)
            if not raw:
                logger.warning(f"  {query_date} 未获取到任何公告（非交易日 / 网络问题？）")
                return None

            # ── Step 2: 标题过滤 ─────────────────────────────────────────
            logger.info("[2/4] 按标题关键词过滤减持公告...")
            reduction_ann = self.fetcher.filter_by_title(raw)
            if not reduction_ann:
                logger.info(f"  {query_date} 过滤后无减持公告")
                return None

            # ── Step 3: 字段提取 ─────────────────────────────────────────
            logger.info(f"[3/4] 提取 {len(reduction_ann)} 条公告的结构化字段...")
            records: List[Dict] = []
            for ann in reduction_ann:
                try:
                    rec = self.extractor.extract(ann)
                    # 可选 PDF 增强
                    if self.enhancer:
                        extra = self.enhancer.enhance_record(rec)
                        rec.update(extra)
                    records.append(rec)
                except Exception as exc:
                    # 降级：保留基本信息，不中断
                    logger.warning(
                        f"  提取失败 [{ann.get('announcementTitle', '')}]: {exc}"
                    )
                    records.append(self._fallback_record(ann, query_date))

            logger.info(f"  ↳ 成功提取 {len(records)} 条记录")

            # ── Step 4: 生成报表 ─────────────────────────────────────────
            logger.info("[4/4] 合并行情数据 → 生成 Excel 报表...")
            out_path = self.reporter.generate(records, query_date)

            logger.info(_banner)
            logger.info(f"  任务完成！共处理 {len(records)} 条减持公告")
            logger.info(f"  报表路径: {out_path}")
            logger.info(_banner)
            return out_path

        except Exception as exc:
            logger.error(f"数据管道执行异常: {exc}", exc_info=True)
            return None

    @staticmethod
    def _fallback_record(ann: Dict, query_date: str) -> Dict:
        """解析失败时的降级记录（保留基本信息）"""
        return {
            "公告日期":          query_date,
            "股票代码":          str(ann.get("secCode", "")).zfill(6),
            "股票简称":          ann.get("secName", ""),
            "减持股东":          "需人工核实",
            "拟减持比例(%)":     "需人工核实",
            "拟减持数量(万股)":  "需人工核实",
            "减持方式":          "需人工核实",
            "股份来源":          "需人工核实",
            "公告标题":          ann.get("announcementTitle", ""),
            "公告原文链接":      build_cninfo_url(ann.get("adjunctUrl", "")),
        }


# ============================================================
# § 10. 自动化调度模块 —— Daemon Scheduler
# ============================================================

def _run_scheduled_job():
    """每日定时任务的执行入口"""
    logger.info(f"定时任务触发 [{datetime.now():%Y-%m-%d %H:%M:%S}]")
    pipeline = ReductionPipeline()
    result = pipeline.run()
    if result:
        logger.info(f"  ✓ 定时任务完成 → {result}")
    else:
        logger.warning("  ✗ 定时任务完成，但未生成报表（无数据或执行失败）")


def run_daemon(schedule_time: str = CONFIG["schedule_time"]):
    """
    以守护进程方式运行，每日指定时间自动执行。

    设计要点:
      - 主循环每 30 秒 poll 一次 schedule
      - KeyboardInterrupt 优雅退出
      - 意外异常不崩溃，记录日志后继续
    """
    logger.info(f"守护进程启动 | 定时时间: 每天 {schedule_time}")
    logger.info("按 Ctrl+C 可停止运行")

    schedule.every().day.at(schedule_time).do(_run_scheduled_job)

    while True:
        try:
            schedule.run_pending()
            time.sleep(30)
        except KeyboardInterrupt:
            logger.info("收到 Ctrl+C，守护进程正常退出")
            break
        except Exception as exc:
            logger.error(f"调度循环异常（将在60s后继续）: {exc}", exc_info=True)
            time.sleep(60)


# ============================================================
# § 11. 命令行入口
# ============================================================

def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="A股减持公告自动化抓取与解析系统",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
运行示例:
  python reduction_scraper.py --mode run                   # 立即抓取今日公告
  python reduction_scraper.py --mode run --date 2025-03-01 # 抓取指定日期
  python reduction_scraper.py --mode daemon                # 守护进程(每日21:00)
  python reduction_scraper.py --mode daemon --time 20:30   # 自定义定时时间
  python reduction_scraper.py --mode run --pdf             # 启用PDF增强解析
        """,
    )
    p.add_argument(
        "--mode",
        choices=["run", "daemon"],
        default="run",
        help="运行模式: run=立即执行, daemon=守护进程",
    )
    p.add_argument(
        "--date",
        type=str,
        default=None,
        metavar="YYYY-MM-DD",
        help="查询日期（仅 run 模式有效），默认为今天",
    )
    p.add_argument(
        "--time",
        type=str,
        default=CONFIG["schedule_time"],
        metavar="HH:MM",
        help=f"定时执行时间（仅 daemon 模式有效），默认 {CONFIG['schedule_time']}",
    )
    p.add_argument(
        "--pdf",
        action="store_true",
        default=False,
        help="启用 PDF 正文增强解析（需安装 pdfplumber）",
    )
    p.add_argument(
        "--output-dir",
        type=str,
        default=CONFIG["output_dir"],
        help=f"Excel 报表输出目录，默认 {CONFIG['output_dir']}",
    )
    return p


def main():
    parser = _build_arg_parser()
    args = parser.parse_args()

    # 动态更新配置
    CONFIG["output_dir"] = args.output_dir

    if args.mode == "run":
        pipeline = ReductionPipeline(enable_pdf_enhance=args.pdf)
        result = pipeline.run(query_date=args.date)
        sys.exit(0 if result else 1)

    elif args.mode == "daemon":
        run_daemon(schedule_time=args.time)


if __name__ == "__main__":
    main()


# ============================================================
# § 附录：字段提取策略详细说明
# ============================================================
"""
┌─────────────────────────────────────────────────────────────────────────────┐
│                   字段提取策略 Field Extraction Strategy                     │
├──────────────────────┬──────────────────────────────────────────────────────┤
│ 字段名               │ 提取逻辑                                              │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 股票代码/简称        │ 直接来自 CNINFO API 的 secCode / secName 字段，       │
│                      │ 无需解析，准确率 100%                                │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 减持股东             │ 层叠正则：                                            │
│                      │ 1. "关于X拟/计划减持" 中提取 X                        │
│                      │ 2. "关于X持股5%以上" 中提取 X                         │
│                      │ 3. 匹配"有限公司/合伙企业/基金"等实体后缀              │
│                      │ 4. 匹配"先生/女士"自然人格式                          │
│                      │ 5. 兜底取"关于"→"拟"之间的文字                        │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 拟减持比例(%)        │ 正则匹配"不超过总股本X%"/"占总股本X%"等6种模式，      │
│                      │ 捕获小数点数字并校验为合法浮点数                      │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 拟减持数量(万股)     │ 正则匹配"不超过X万股"/"X万股"等5种模式，              │
│                      │ 若单位为"股"则自动÷10000转为万股                     │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 减持方式             │ 关键词字典匹配，可同时命中多种方式（如                │
│                      │ "大宗交易或集中竞价"→ "大宗交易、集中竞价"）          │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 股份来源             │ 关键词字典匹配，覆盖IPO前/定增/激励/员工持股等场景    │
│                      │ 通常需 PDF 正文才能精确提取，启用 --pdf 参数可增强    │
├──────────────────────┼──────────────────────────────────────────────────────┤
│ 无法提取的字段       │ 填入"需人工核实"（Excel中黄色高亮），                 │
│                      │ 保留"公告原文链接"供人工复核，全程不中断              │
└──────────────────────┴──────────────────────────────────────────────────────┘

数据来源优先级:
  Level 1 (默认): 巨潮资讯 API 标题文字 → 正则提取
  Level 2 (--pdf): 下载 PDF + pdfplumber 解析正文 → 补充缺失字段
  Level 3 (未来可扩展): 调用大模型 API (GPT/Claude) 做语义抽取

巨潮资讯 API 补充说明:
  - 接口地址: POST http://www.cninfo.com.cn/new/hisAnnouncement/query
  - 认证方式: 无需登录，但需要正确 Referer 和 Content-Type 头
  - 限速建议: 请求间隔 ≥ 1.5秒，max_pages 设为 15（覆盖约450条/关键词）
  - adjunctUrl 格式: "YYYY-MM-DD/filename.PDF"
  - 完整 PDF URL: http://static.cninfo.com.cn/finalpage/{adjunctUrl}
"""
